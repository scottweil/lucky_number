from openpyxl import load_workbook
import random
import time
import csv

win_record_load = load_workbook("당첨번호.xlsx", data_only=True)

load_ws = win_record_load["Sheet1"]

get_cells = load_ws["A1":"F925"]

print(f"There are {load_ws.max_row} game samples for now")
time.sleep(1)

win_num_list = []
for row in get_cells:
    for cell in row:
        win_num_list.append(str(cell.value))

print("you are gonna get lucky numbers for 5 games")
time.sleep(1)

for n in range(5):
    print(5 - n)
    time.sleep(1)

expectation = []
expectation_list = []
for n in range(5):
    while True:
        expectation.append(random.choice(win_num_list))
        expectation_set = set(expectation)
        if len(expectation_set) == 6:
            expectation_list.append(expectation_set)
            expectation.clear()
            break

fw = open("lucky_number.csv", "w", encoding="utf-8", newline="")
wr = csv.writer(fw)
wr.writerows(expectation_list)
fw.close()

print("good luck!")

fr = open("lucky_number.csv", "r", encoding="utf-8")
rd = csv.reader(fr)
for i in rd:
    print(i)

fr.close()
input()